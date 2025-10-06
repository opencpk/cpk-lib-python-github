# pylint: disable=too-many-positional-arguments

"""Export functionality for backup data"""

import csv
import json
from dataclasses import asdict

from openpyxl import Workbook  # pylint: disable=import-error
from openpyxl.styles import (  # pylint: disable=import-error
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)

from .models import OrganizationBackup


def export_to_json(backup: OrganizationBackup, filename: str):
    """Export backup to JSON format"""
    backup_dict = asdict(backup)

    with open(filename, "w", encoding="utf-8") as f:
        json.dump(backup_dict, f, indent=2, ensure_ascii=False)

    print(f"📄 JSON backup saved to: {filename}")


def export_to_csv(backup: OrganizationBackup, filename: str):
    """Export backup to CSV format (original single file)"""
    with open(filename, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        # Write header
        writer.writerow(
            [
                "Username",
                "User ID",
                "Email",
                "Role",
                "Team Name",
                "Team Slug",
                "Team Role",
                "Team Privacy",
                "Team Repositories",
            ]
        )

        # Write data
        for user in backup.users:
            base_row = [user.username, user.user_id, user.email, user.role]

            # Team membership rows
            for team in user.teams:
                writer.writerow(
                    base_row
                    + [team.name, team.slug, team.role, team.privacy, "; ".join(team.repositories)]
                )

            # User with no teams
            if not user.teams:
                writer.writerow(base_row + ["", "", "", "", ""])

    print(f"📊 CSV backup saved to: {filename}")


def export_to_multiple_csvs(backup: OrganizationBackup, base_filename: str):
    """Export backup to multiple focused CSV files"""
    base_name = base_filename.replace(".json", "")

    print("\n📋 Exporting to multiple CSV files...")

    # 1. Teams Overview
    teams_file = f"{base_name}_teams_overview.csv"
    with open(teams_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            [
                "Team Name",
                "Team Slug",
                "Privacy",
                "Description",
                "Member Count",
                "Repository Count",
                "Repository List",
            ]
        )

        # Build team data from user memberships
        team_data = {}
        for user in backup.users:
            for team in user.teams:
                if team.slug not in team_data:
                    team_data[team.slug] = {
                        "name": team.name,
                        "privacy": team.privacy,
                        "description": team.description or "",
                        "members": set(),
                        "repositories": set(team.repositories),
                    }
                team_data[team.slug]["members"].add(user.username)

        # Write team overview data sorted by team name
        for slug, data in sorted(team_data.items(), key=lambda x: x[1]["name"].lower()):
            repo_list = "; ".join(sorted(data["repositories"]))
            writer.writerow(
                [
                    data["name"],
                    slug,
                    data["privacy"],
                    data["description"],
                    len(data["members"]),
                    len(data["repositories"]),
                    repo_list,
                ]
            )

    # 2. Team Memberships
    memberships_file = f"{base_name}_team_memberships.csv"
    with open(memberships_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(
            ["Team Name", "Team Slug", "Username", "User Email", "Team Role", "Org Role"]
        )

        # Collect all memberships and sort by team name, then username
        memberships = []
        for user in backup.users:
            for team in user.teams:
                memberships.append(
                    {
                        "team_name": team.name,
                        "team_slug": team.slug,
                        "username": user.username,
                        "email": user.email or "",
                        "team_role": team.role,
                        "org_role": user.role,
                    }
                )

        # Sort by team name, then username
        memberships.sort(key=lambda x: (x["team_name"].lower(), x["username"].lower()))

        for membership in memberships:
            writer.writerow(
                [
                    membership["team_name"],
                    membership["team_slug"],
                    membership["username"],
                    membership["email"],
                    membership["team_role"],
                    membership["org_role"],
                ]
            )

    # 3. Team Repository Access
    team_repos_file = f"{base_name}_team_repositories.csv"
    with open(team_repos_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Team Name", "Team Slug", "Repository", "Repository Org/Name"])

        # Collect unique team-repo combinations
        team_repo_access = set()
        for user in backup.users:
            for team in user.teams:
                for repo in team.repositories:
                    team_repo_access.add((team.name, team.slug, repo))

        # Sort by team name, then repo name
        for team_name, team_slug, repo in sorted(team_repo_access):
            repo_parts = repo.split("/")
            repo_short = repo_parts[-1] if repo_parts else repo
            writer.writerow([team_name, team_slug, repo_short, repo])

    # 4. Users Summary
    users_file = f"{base_name}_users_summary.csv"
    with open(users_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Username", "User ID", "Email", "Org Role", "Team Count", "Team List"])

        # Sort users by username
        for user in sorted(backup.users, key=lambda x: x.username.lower()):
            team_list = "; ".join([team.name for team in user.teams])
            writer.writerow(
                [
                    user.username,
                    user.user_id,
                    user.email or "",
                    user.role,
                    len(user.teams),
                    team_list,
                ]
            )

    # 5. Users Without Teams
    no_teams_file = f"{base_name}_users_without_teams.csv"
    with open(no_teams_file, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Username", "User ID", "Email", "Org Role"])

        users_without_teams = [user for user in backup.users if not user.teams]
        for user in sorted(users_without_teams, key=lambda x: x.username.lower()):
            writer.writerow([user.username, user.user_id, user.email or "", user.role])

    print("📊 Multiple CSV files exported:")
    print(f"   📋 {teams_file} - Teams overview ({len(team_data)} teams)")
    print(f"   👥 {memberships_file} - Team memberships ({len(memberships)} memberships)")
    print(
        f"   📚 {team_repos_file} - Team repository access ({len(team_repo_access)} team-repo pairs)"
    )
    print(f"   👤 {users_file} - Users summary ({len(backup.users)} users)")
    print(f"   ❌ {no_teams_file} - Users without teams ({len(users_without_teams)} users)")


def export_to_excel(backup: OrganizationBackup, base_filename: str):
    """Export backup to Excel with rich formatting"""
    # Check openpyxl at runtime instead of import time

    excel_filename = base_filename.replace(".json", ".xlsx")
    wb = Workbook()

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    team_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    repo_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Remove default sheet
    wb.remove(wb.active)

    # 1. Teams Overview Sheet
    _create_teams_overview_sheet(
        wb, backup, header_font, header_fill, center_alignment, left_alignment, border, alt_row_fill
    )

    # 2. Team Memberships Sheet
    _create_team_memberships_sheet(
        wb, backup, team_header_fill, center_alignment, left_alignment, border
    )

    # 3. Team Repositories Sheet
    _create_team_repositories_sheet(
        wb,
        backup,
        team_header_fill,
        repo_header_fill,
        center_alignment,
        left_alignment,
        border,
        alt_row_fill,
    )

    # 4. Users Summary Sheet
    _create_users_summary_sheet(
        wb, backup, header_font, header_fill, center_alignment, left_alignment, border, alt_row_fill
    )

    # 5. Users Without Teams Sheet
    _create_users_without_teams_sheet(
        wb, backup, header_font, header_fill, center_alignment, left_alignment, border, alt_row_fill
    )

    # Save the workbook
    wb.save(excel_filename)

    print(f"📊 Excel file exported: {excel_filename}")
    print("   📋 Teams Overview - Summary of all teams")
    print("   👥 Team Memberships - Grouped by team with rich formatting")
    print("   📚 Team Repositories - Grouped by team with repository details")
    print("   👤 Users Summary - All users with team information")
    print("   ❌ Users Without Teams - Users not in any teams")


def _create_teams_overview_sheet(
    wb, backup, header_font, header_fill, center_alignment, left_alignment, border, alt_row_fill
):
    """Create Teams Overview sheet"""
    ws_teams = wb.create_sheet("Teams Overview")

    # Headers
    headers = [
        "Team Name",
        "Team Slug",
        "Privacy",
        "Description",
        "Member Count",
        "Repository Count",
        "Repository List",
    ]
    for col, header in enumerate(headers, 1):
        cell = ws_teams.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    # Build team data
    team_data = {}
    for user in backup.users:
        for team in user.teams:
            if team.slug not in team_data:
                team_data[team.slug] = {
                    "name": team.name,
                    "privacy": team.privacy,
                    "description": team.description or "",
                    "members": set(),
                    "repositories": set(team.repositories),
                }
            team_data[team.slug]["members"].add(user.username)

    # Data rows
    for row, (slug, data) in enumerate(
        sorted(team_data.items(), key=lambda x: x[1]["name"].lower()), 2
    ):
        repo_list = "; ".join(sorted(data["repositories"]))
        values = [
            data["name"],
            slug,
            data["privacy"],
            data["description"],
            len(data["members"]),
            len(data["repositories"]),
            repo_list,
        ]

        for col, value in enumerate(values, 1):
            cell = ws_teams.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = left_alignment if col in [1, 3, 7] else center_alignment
            if row % 2 == 0:
                cell.fill = alt_row_fill

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws_teams.column_dimensions[ws_teams.cell(row=1, column=col).column_letter].width = 15
    ws_teams.column_dimensions["A"].width = 25
    ws_teams.column_dimensions["D"].width = 30
    ws_teams.column_dimensions["G"].width = 50


def _create_team_memberships_sheet(
    wb, backup, team_header_fill, center_alignment, left_alignment, border
):
    """Create Team Memberships sheet"""

    ws_memberships = wb.create_sheet("Team Memberships")

    # Group memberships by team
    team_memberships = {}
    for user in backup.users:
        for team in user.teams:
            if team.slug not in team_memberships:
                team_memberships[team.slug] = {"name": team.name, "members": []}
            team_memberships[team.slug]["members"].append(
                {
                    "username": user.username,
                    "email": user.email or "",
                    "team_role": team.role,
                    "org_role": user.role,
                }
            )

    current_row = 1

    for _, team_info in sorted(team_memberships.items(), key=lambda x: x[1]["name"].lower()):
        # Team header
        ws_memberships.merge_cells(f"A{current_row}:E{current_row}")
        team_cell = ws_memberships.cell(row=current_row, column=1, value=team_info["name"])
        team_cell.font = Font(bold=True, color="FFFFFF", size=14)
        team_cell.fill = team_header_fill
        team_cell.alignment = center_alignment
        team_cell.border = border

        # Apply border to merged cells
        for col in range(1, 6):
            ws_memberships.cell(row=current_row, column=col).border = border
            ws_memberships.cell(row=current_row, column=col).fill = team_header_fill

        current_row += 1

        # Member count
        member_count_cell = ws_memberships.cell(
            row=current_row, column=1, value=f"{len(team_info['members'])} members"
        )
        member_count_cell.font = Font(italic=True)
        member_count_cell.alignment = left_alignment
        current_row += 1

        # Column headers for this team
        member_headers = ["Username", "User Email", "Team Role", "Org Role"]
        for col, header in enumerate(member_headers, 1):
            cell = ws_memberships.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            cell.alignment = center_alignment
            cell.border = border
        current_row += 1

        # Member data
        for member in sorted(team_info["members"], key=lambda x: x["username"].lower()):
            values = [member["username"], member["email"], member["team_role"], member["org_role"]]

            for col, value in enumerate(values, 1):
                cell = ws_memberships.cell(row=current_row, column=col, value=value)
                cell.border = border
                cell.alignment = left_alignment if col in [1, 2] else center_alignment

            current_row += 1

        current_row += 1

    # Auto-adjust column widths
    ws_memberships.column_dimensions["A"].width = 20
    ws_memberships.column_dimensions["B"].width = 25
    ws_memberships.column_dimensions["C"].width = 15
    ws_memberships.column_dimensions["D"].width = 15


def _create_team_repositories_sheet(
    wb,
    backup,
    team_header_fill,
    repo_header_fill,
    center_alignment,
    left_alignment,
    border,
    alt_row_fill,
):
    """Create Team Repositories sheet"""

    ws_repos = wb.create_sheet("Team Repositories")

    # Group repositories by team
    team_repositories = {}
    for user in backup.users:
        for team in user.teams:
            if team.slug not in team_repositories:
                team_repositories[team.slug] = {
                    "name": team.name,
                    "repositories": set(team.repositories),
                }
            else:
                team_repositories[team.slug]["repositories"].update(team.repositories)

    current_row = 1

    for _, team_info in sorted(team_repositories.items(), key=lambda x: x[1]["name"].lower()):
        if not team_info["repositories"]:
            continue

        # Team header
        ws_repos.merge_cells(f"A{current_row}:D{current_row}")
        team_cell = ws_repos.cell(row=current_row, column=1, value=team_info["name"])
        team_cell.font = Font(bold=True, color="FFFFFF", size=14)
        team_cell.fill = team_header_fill
        team_cell.alignment = center_alignment
        team_cell.border = border

        # Apply border to merged cells
        for col in range(1, 5):
            ws_repos.cell(row=current_row, column=col).border = border
            ws_repos.cell(row=current_row, column=col).fill = team_header_fill

        current_row += 1

        # Repository count
        repo_count_cell = ws_repos.cell(
            row=current_row, column=1, value=f"{len(team_info['repositories'])} repositories"
        )
        repo_count_cell.font = Font(italic=True)
        repo_count_cell.alignment = left_alignment
        current_row += 1

        # Column headers for this team's repositories
        repo_headers = ["Repository Name", "Full Repository Path", "Organization", "Project"]
        for col, header in enumerate(repo_headers, 1):
            cell = ws_repos.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = repo_header_fill
            cell.alignment = center_alignment
            cell.border = border
        current_row += 1

        # Repository data
        for repo in sorted(team_info["repositories"]):
            repo_parts = repo.split("/")
            org_name = repo_parts[0] if len(repo_parts) > 1 else ""
            repo_name = repo_parts[-1] if repo_parts else repo

            values = [repo_name, repo, org_name, repo_name]

            for col, value in enumerate(values, 1):
                cell = ws_repos.cell(row=current_row, column=col, value=value)
                cell.border = border
                cell.alignment = left_alignment
                if current_row % 2 == 0:
                    cell.fill = alt_row_fill

            current_row += 1

        current_row += 1

    # Auto-adjust column widths
    ws_repos.column_dimensions["A"].width = 30
    ws_repos.column_dimensions["B"].width = 50
    ws_repos.column_dimensions["C"].width = 20
    ws_repos.column_dimensions["D"].width = 30


# pylint: disable=too-many-positional-arguments
def _create_users_summary_sheet(
    wb, backup, header_font, header_fill, center_alignment, left_alignment, border, alt_row_fill
):
    """Create Users Summary sheet"""
    ws_users = wb.create_sheet("Users Summary")

    headers = ["Username", "User ID", "Email", "Org Role", "Team Count", "Team List"]
    for col, header in enumerate(headers, 1):
        cell = ws_users.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    # Data rows
    for row, user in enumerate(sorted(backup.users, key=lambda x: x.username.lower()), 2):
        team_list = "; ".join([team.name for team in user.teams])
        values = [
            user.username,
            user.user_id,
            user.email or "",
            user.role,
            len(user.teams),
            team_list,
        ]

        for col, value in enumerate(values, 1):
            cell = ws_users.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = left_alignment if col in [1, 3, 6] else center_alignment
            if row % 2 == 0:
                cell.fill = alt_row_fill

    # Auto-adjust column widths
    ws_users.column_dimensions["A"].width = 20
    ws_users.column_dimensions["B"].width = 12
    ws_users.column_dimensions["C"].width = 25
    ws_users.column_dimensions["D"].width = 12
    ws_users.column_dimensions["E"].width = 12
    ws_users.column_dimensions["F"].width = 50


# pylint: disable=too-many-positional-arguments
def _create_users_without_teams_sheet(
    wb, backup, header_font, header_fill, center_alignment, left_alignment, border, alt_row_fill
):
    """Create Users Without Teams sheet"""
    ws_no_teams = wb.create_sheet("Users Without Teams")

    headers = ["Username", "User ID", "Email", "Org Role"]
    for col, header in enumerate(headers, 1):
        cell = ws_no_teams.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    users_without_teams = [user for user in backup.users if not user.teams]
    for row, user in enumerate(sorted(users_without_teams, key=lambda x: x.username.lower()), 2):
        values = [user.username, user.user_id, user.email or "", user.role]

        for col, value in enumerate(values, 1):
            cell = ws_no_teams.cell(row=row, column=col, value=value)
            cell.border = border
            cell.alignment = left_alignment if col in [1, 3] else center_alignment
            if row % 2 == 0:
                cell.fill = alt_row_fill

    # Auto-adjust column widths
    ws_no_teams.column_dimensions["A"].width = 20
    ws_no_teams.column_dimensions["B"].width = 12
    ws_no_teams.column_dimensions["C"].width = 25
    ws_no_teams.column_dimensions["D"].width = 12


def export_to_structured_json(backup: OrganizationBackup, filename: str):
    """Export to clean, well-structured JSON"""

    # Build team-centric structure
    teams_structure = {}

    for user in backup.users:
        for team in user.teams:
            if team.slug not in teams_structure:
                teams_structure[team.slug] = {
                    "name": team.name,
                    "slug": team.slug,
                    "description": team.description,
                    "privacy": team.privacy,
                    "repositories": team.repositories,
                    "members": [],
                }

            teams_structure[team.slug]["members"].append(
                {
                    "username": user.username,
                    "user_id": user.user_id,
                    "email": user.email,
                    "role_in_team": team.role,
                    "org_role": user.role,
                }
            )

    # Build users without teams
    users_without_teams = []
    for user in backup.users:
        if not user.teams:
            users_without_teams.append(
                {
                    "username": user.username,
                    "user_id": user.user_id,
                    "email": user.email,
                    "org_role": user.role,
                }
            )

    # Build final structure
    structured_data = {
        "organization": backup.org_name,
        "backup_timestamp": backup.backup_timestamp,
        "backup_type": backup.backup_type,
        "summary": backup.summary,
        "teams": sorted(list(teams_structure.values()), key=lambda x: x["name"].lower()),
        "users_without_teams": sorted(users_without_teams, key=lambda x: x["username"].lower()),
    }

    with open(filename, "w", encoding="utf-8") as f:
        json.dump(structured_data, f, indent=2, ensure_ascii=False)

    print(f"📄 Structured JSON saved to: {filename}")
